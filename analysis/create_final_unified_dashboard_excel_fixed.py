import json
import pandas as pd
import numpy as np
import optuna
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
import optuna_dashboard
import webbrowser
import time
import subprocess
import sys
import sqlite3

def load_studies_from_unified_db(experiment_name):
    """통합 DB에서 모든 Optuna study 로드"""
    model_types = ['DCNV2', 'DCNV2_FUXICTR', 'CUSTOM_FOCAL_DL', 'CUSTOM_NN_TORCH', 'RF']
    studies = {}
    
    # 실험별 DB 경로 자동 구성
    db_path = f'optuna_studies/{experiment_name}/all_studies.db'
    unified_db_path = f'sqlite:///{db_path}'
    
    print(f"🔍 DB 경로: {db_path}")
    
    for model_type in model_types:
        try:
            study = optuna.load_study(
                study_name=f'{model_type}_hpo_study',
                storage=unified_db_path
            )
            studies[model_type] = study
            print(f"✅ {model_type} study 로드 완료 (통합 DB)")
        except Exception as e:
            print(f"❌ {model_type} study 로드 실패: {e}")
    
    return studies

def analyze_parameter_importance(studies):
    """파라미터 중요도 분석"""
    importance_results = {}
    
    for model_type, study in studies.items():
        try:
            importance = optuna.importance.get_param_importances(study)
            importance_results[model_type] = importance
            print(f"✅ {model_type} 파라미터 중요도 계산 완료")
        except Exception as e:
            print(f"❌ {model_type} 파라미터 중요도 계산 실패: {e}")
            importance_results[model_type] = {}
    
    return importance_results

def analyze_optimization_history(studies):
    """최적화 과정 분석"""
    history_results = {}
    
    for model_type, study in studies.items():
        try:
            trials = study.trials
            successful_trials = [t for t in trials if t.state == optuna.trial.TrialState.COMPLETE]
            
            if len(successful_trials) >= 2:
                values = [t.value for t in successful_trials]
                best_value = study.best_value
                mean_value = np.mean(values)
                std_value = np.std(values)
                
                # 수렴성 분석
                if len(values) >= 5:
                    recent_values = values[-5:]
                    improvement = recent_values[-1] - recent_values[0]
                    
                    if improvement > 0.01:
                        convergence = "개선 중"
                    elif abs(improvement) < 0.005:
                        convergence = "수렴됨"
                    else:
                        convergence = "불안정"
                else:
                    convergence = "데이터 부족"
                
                history_results[model_type] = {
                    'best_value': best_value,
                    'mean_value': mean_value,
                    'std_value': std_value,
                    'min_value': min(values),
                    'max_value': max(values),
                    'convergence': convergence,
                    'success_rate': len(successful_trials)/len(trials)*100,
                    'total_trials': len(successful_trials)
                }
                
        except Exception as e:
            print(f"❌ {model_type} 최적화 과정 분석 실패: {e}")
    
    return history_results

def generate_custom_recommendations(studies, importance_results, history_results):
    """사용자 지정 형식의 권장사항 생성"""
    recommendations_text = """
    <div class="criteria-section">
        <h4>📋 권장사항 기준 설명</h4>
        <ul>
            <li><strong>변동성 분석 (표준편차 > 0.05):</strong> 성능이 불안정하면 더 많은 trial 필요</li>
            <li><strong>수렴성 분석 (최근 5개 trial):</strong> 개선 추세, 수렴, 불안정 상태 판단</li>
            <li><strong>파라미터 중요도:</strong> 가장 영향력 있는 파라미터 우선 탐색</li>
            <li><strong>상관관계 분석:</strong> 강한 상관관계가 있는 파라미터는 함께 조정</li>
        </ul>
    </div>
    """
    
    # 각 모델별 권장사항
    for model_type in studies.keys():
        history = history_results.get(model_type, {})
        importance = importance_results.get(model_type, {})
        
        if history:
            recommendations_text += f"""
            <h4>🎯 {model_type}</h4>
            <ul>
            """
            
            # 변동성 분석
            std_value = history.get('std_value', 0)
            if std_value > 0.05:
                recommendations_text += f"<li>📈 더 많은 trial 필요 (높은 변동성: {std_value:.4f})</li>"
            else:
                recommendations_text += "<li>✅ 현재 설정으로 충분히 최적화됨</li>"
            
            # 수렴성 분석
            convergence = history.get('convergence', '')
            if convergence == "불안정":
                recommendations_text += "<li>🔧 더 세밀한 파라미터 탐색 필요</li>"
            elif convergence == "개선 중":
                recommendations_text += "<li>📊 더 많은 trial로 추가 개선 가능</li>"
            
            # 파라미터 중요도
            if importance:
                top_param = max(importance, key=importance.get)
                top_importance = importance[top_param]
                recommendations_text += f"<li>🎯 가장 중요한 파라미터: {top_param} (중요도: {top_importance:.4f})</li>"
            
            recommendations_text += "</ul>"
    
    return recommendations_text

def create_excel_report_unified_db(studies, importance_results, history_results, experiment_name="experiment"):
    """고급 분석 리포트 엑셀 생성 (통합 DB)"""
    print("=== 고급 분석 리포트 엑셀 생성 (통합 DB) ===")
    
    # 실험별 결과 폴더 생성
    experiment_results_dir = f"results/{experiment_name}"
    os.makedirs(experiment_results_dir, exist_ok=True)
    
    # 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{experiment_results_dir}/optuna_advanced_report_{experiment_name}_{timestamp}.xlsx"
    
    # 워크북 생성
    wb = Workbook()
    
    # 기본 시트 제거
    wb.remove(wb.active)
    
    # 1. 요약 시트
    create_summary_sheet(wb, studies, history_results, experiment_name)
    
    # 2. 파라미터 중요도 시트
    create_importance_sheet(wb, importance_results)
    
    # 3. 최적화 과정 시트
    create_optimization_sheet(wb, history_results)
    
    # 4. 각 모델별 상세 시트
    for model_type in studies.keys():
        create_model_detail_sheet(wb, model_type, studies[model_type], 
                                importance_results.get(model_type, {}),
                                history_results.get(model_type, {}))
    
    # 5. 권장사항 시트
    create_recommendations_sheet(wb, studies, importance_results, history_results)
    
    # 파일 저장
    wb.save(filename)
    
    print(f"✅ 고급 분석 리포트가 '{filename}'에 저장되었습니다!")
    return filename

def create_summary_sheet(wb, studies, history_results, experiment_name):
    """요약 시트 생성"""
    ws = wb.create_sheet("📊 실험 요약")
    
    # 제목
    ws['A1'] = f"🎯 Optuna HPO 고급 분석 리포트 - {experiment_name}"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    # 생성 시간
    ws['A2'] = f"생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(size=10, italic=True)
    
    # 실험 정보
    ws['A3'] = f"실험 이름: {experiment_name}"
    ws['A3'].font = Font(size=10, bold=True, color="0066CC")
    
    # DB 정보
    ws['A4'] = f"DB 파일: optuna_studies/{experiment_name}/all_studies.db"
    ws['A4'].font = Font(size=10, italic=True, color="0066CC")
    
    # 요약 통계
    row = 5
    ws[f'A{row}'] = "📈 전체 실험 통계"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row += 1
    total_trials = sum([history.get('total_trials', 0) for history in history_results.values()])
    best_performance = max([history.get('best_value', 0) for history in history_results.values()]) if history_results else 0
    avg_performance = np.mean([history.get('mean_value', 0) for history in history_results.values()]) if history_results else 0
    
    summary_data = [
        ["총 Trial 수", total_trials],
        ["최고 성능", f"{best_performance:.4f}"],
        ["평균 성능", f"{avg_performance:.4f}"],
        ["분석된 모델 수", len(studies)]
    ]
    
    for i, (label, value) in enumerate(summary_data):
        ws[f'A{row+i}'] = label
        ws[f'B{row+i}'] = value
        ws[f'A{row+i}'].font = Font(bold=True)
    
    # 모델별 성능 비교
    row += len(summary_data) + 2
    ws[f'A{row}'] = "🏆 모델별 성능 비교"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row += 1
    headers = ["모델", "최고 성능", "평균 성능", "표준편차", "수렴 상태", "성공률(%)"]
    for i, header in enumerate(headers):
        ws.cell(row=row, column=i+1, value=header).font = Font(bold=True)
        ws.cell(row=row, column=i+1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    row += 1
    for model_type, history in history_results.items():
        if history:
            data = [
                model_type,
                f"{history['best_value']:.4f}",
                f"{history['mean_value']:.4f}",
                f"{history['std_value']:.4f}",
                history['convergence'],
                f"{history['success_rate']:.1f}%"
            ]
            for i, value in enumerate(data):
                ws.cell(row=row, column=i+1, value=value)
            row += 1

def create_importance_sheet(wb, importance_results):
    """파라미터 중요도 시트 생성"""
    ws = wb.create_sheet("🎯 파라미터 중요도")
    
    # 제목
    ws['A1'] = "파라미터 중요도 분석"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:D1')
    
    row = 3
    for model_type, importance in importance_results.items():
        if importance:
            ws[f'A{row}'] = f"🎯 {model_type}"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            row += 1
            ws[f'A{row}'] = "파라미터"
            ws[f'B{row}'] = "중요도"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            
            row += 1
            for param, imp in sorted(importance.items(), key=lambda x: x[1], reverse=True):
                ws[f'A{row}'] = param
                ws[f'B{row}'] = f"{imp:.4f}"
                
                # 조건부 서식 (중요도에 따른 색상)
                if imp > 0.5:
                    ws[f'B{row}'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                elif imp > 0.3:
                    ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                
                row += 1
            
            row += 2

def create_optimization_sheet(wb, history_results):
    """최적화 과정 시트 생성"""
    ws = wb.create_sheet("📈 최적화 과정")
    
    # 제목
    ws['A1'] = "최적화 과정 분석"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    row = 3
    headers = ["모델", "최고값", "평균값", "표준편차", "최소값", "최대값", "수렴상태"]
    for i, header in enumerate(headers):
        ws.cell(row=row, column=i+1, value=header).font = Font(bold=True)
        ws.cell(row=row, column=i+1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    row += 1
    for model_type, history in history_results.items():
        if history:
            data = [
                model_type,
                f"{history['best_value']:.4f}",
                f"{history['mean_value']:.4f}",
                f"{history['std_value']:.4f}",
                f"{history['min_value']:.4f}",
                f"{history['max_value']:.4f}",
                history['convergence']
            ]
            for i, value in enumerate(data):
                cell = ws.cell(row=row, column=i+1, value=value)
                
                # 수렴상태에 따른 조건부 서식
                if i == 6:  # 수렴상태 컬럼
                    if value == "수렴됨":
                        cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
                    elif value == "불안정":
                        cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
                    elif value == "개선 중":
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            row += 1

def create_model_detail_sheet(wb, model_type, study, importance, history):
    """모델별 상세 시트 생성"""
    ws = wb.create_sheet(f"📋 {model_type}_상세")
    
    # 제목
    ws['A1'] = f"{model_type} 상세 분석"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:E1')
    
    # Trial 데이터
    trials = [t for t in study.trials if t.state == optuna.trial.TrialState.COMPLETE]
    if trials:
        row = 3
        ws[f'A{row}'] = "Trial 번호"
        ws[f'B{row}'] = "성능"
        ws[f'C{row}'] = "상태"
        
        # 파라미터 헤더
        all_params = set()
        for trial in trials:
            all_params.update(trial.params.keys())
        
        param_list = sorted(list(all_params))
        for i, param in enumerate(param_list):
            ws.cell(row=row, column=4+i, value=param).font = Font(bold=True)
        
        # 헤더 서식
        for col in range(1, 4 + len(param_list)):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.cell(row=row, column=col).font = Font(bold=True)
        
        row += 1
        for trial in trials:
            ws[f'A{row}'] = trial.number
            ws[f'B{row}'] = f"{trial.value:.4f}"
            ws[f'C{row}'] = trial.state.name
            
            # 파라미터 값
            for i, param in enumerate(param_list):
                value = trial.params.get(param, "")
                ws.cell(row=row, column=4+i, value=str(value))
            
            # 최고 성능 강조
            if trial.value == study.best_value:
                for col in range(1, 4 + len(param_list)):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            
            row += 1

def create_recommendations_sheet(wb, studies, importance_results, history_results):
    """권장사항 시트 생성"""
    ws = wb.create_sheet("💡 권장사항")
    
    # 제목
    ws['A1'] = "다음 실험을 위한 권장사항"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:D1')
    
    row = 3
    ws[f'A{row}'] = "📋 권장사항 기준 설명"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    criteria = [
        "변동성 분석 (표준편차 > 0.05): 성능이 불안정하면 더 많은 trial 필요",
        "수렴성 분석 (최근 5개 trial): 개선 추세, 수렴, 불안정 상태 판단",
        "파라미터 중요도: 가장 영향력 있는 파라미터 우선 탐색",
        "상관관계 분석: 강한 상관관계가 있는 파라미터는 함께 조정"
    ]
    
    row += 1
    for criterion in criteria:
        ws[f'A{row}'] = f"• {criterion}"
        row += 1
    
    row += 1
    
    # 각 모델별 권장사항
    for model_type in history_results.keys():
        history = history_results.get(model_type, {})
        importance = importance_results.get(model_type, {})
        
        if history:
            ws[f'A{row}'] = f"🎯 {model_type}"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            # 변동성 분석
            std_value = history.get('std_value', 0)
            if std_value > 0.05:
                ws[f'A{row}'] = f"📈 더 많은 trial 필요 (높은 변동성: {std_value:.4f})"
            else:
                ws[f'A{row}'] = "✅ 현재 설정으로 충분히 최적화됨"
            row += 1
            
            # 수렴성 분석
            convergence = history.get('convergence', '')
            if convergence == "불안정":
                ws[f'A{row}'] = "🔧 더 세밀한 파라미터 탐색 필요"
                row += 1
            elif convergence == "개선 중":
                ws[f'A{row}'] = "📊 더 많은 trial로 추가 개선 가능"
                row += 1
            
            # 파라미터 중요도
            if importance:
                top_param = max(importance, key=importance.get)
                top_importance = importance[top_param]
                ws[f'A{row}'] = f"🎯 가장 중요한 파라미터: {top_param} (중요도: {top_importance:.4f})"
                row += 1
            
            row += 1

def safe_json_dumps(obj):
    """안전한 JSON 직렬화"""
    return json.dumps(obj, ensure_ascii=False, default=str)

def create_unified_db_dashboard_with_fixed_correlation(studies, experiment_name="experiment"):
    """통합 DB 대시보드 생성 (상관관계 차트 개선)"""
    print("=== 통합 DB 대시보드 생성 (상관관계 차트 개선) ===")
    
    # 실험별 결과 폴더 생성
    experiment_results_dir = f"results/{experiment_name}"
    os.makedirs(experiment_results_dir, exist_ok=True)
    
    # 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{experiment_results_dir}/optuna_unified_dashboard_correlation_fixed_{experiment_name}_{timestamp}.html"
    
    # 데이터 준비 (상관관계 개선)
    all_data = {}
    for model_type, study in studies.items():
        trials = [t for t in study.trials if t.state == optuna.trial.TrialState.COMPLETE]
        if trials:
            # 기본 데이터
            trial_data = []
            for trial in trials:
                data = {
                    'trial_number': trial.number,
                    'value': trial.value,
                    'state': trial.state.name
                }
                data.update(trial.params)
                trial_data.append(data)
            
            # 파라미터 중요도
            importance = importance_results.get(model_type, {})
            
            # 파라미터 상관관계 (개선된 버전)
            numeric_params = {}
            for trial in trials:
                for param, value in trial.params.items():
                    if isinstance(value, (int, float)):
                        if param not in numeric_params:
                            numeric_params[param] = []
                        numeric_params[param].append(value)
            
            correlation_data = {}
            if len(numeric_params) >= 2:
                param_names = list(numeric_params.keys())
                for i, param1 in enumerate(param_names):
                    for j, param2 in enumerate(param_names):
                        if i < j and len(numeric_params[param1]) == len(numeric_params[param2]):  # 중복 제거
                            corr = np.corrcoef(numeric_params[param1], numeric_params[param2])[0, 1]
                            if not np.isnan(corr):
                                # 더 명확한 라벨링
                                pair_name = f"{param1} ↔ {param2}"
                                correlation_data[pair_name] = corr
            
            all_data[model_type] = {
                'trials': trial_data,
                'importance': importance,
                'correlation': correlation_data,
                'best_value': study.best_value,
                'param_names': list(set().union(*[trial.params.keys() for trial in trials]))
            }
    
    # HTML 시작
    html_content = """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>🎯 Optuna HPO 통합 DB 대시보드 (상관관계 개선)</title>
        <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        <style>
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                margin: 0;
                padding: 20px;
                background-color: #f5f5f5;
            }
            .container {
                max-width: 1800px;
                margin: 0 auto;
                background-color: white;
                border-radius: 10px;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                overflow: hidden;
            }
            .header {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 30px;
                text-align: center;
            }
            .header h1 {
                margin: 0;
                font-size: 2.5em;
                font-weight: 300;
            }
            .header p {
                margin: 10px 0 0 0;
                opacity: 0.9;
                font-size: 1.1em;
            }
            .content {
                padding: 30px;
            }
            .section {
                margin-bottom: 40px;
                background-color: #fafafa;
                border-radius: 8px;
                padding: 25px;
                border-left: 4px solid #667eea;
            }
            .section h2 {
                color: #333;
                margin-top: 0;
                font-size: 1.8em;
                border-bottom: 2px solid #eee;
                padding-bottom: 10px;
            }
            .model-section {
                background-color: white;
                border-radius: 8px;
                padding: 25px;
                margin-bottom: 30px;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            }
            .model-section h3 {
                color: #667eea;
                margin-top: 0;
                font-size: 1.5em;
                border-bottom: 2px solid #eee;
                padding-bottom: 10px;
            }
            .chart-row {
                display: flex;
                gap: 20px;
                margin: 20px 0;
                align-items: flex-start;
            }
            .chart-container {
                flex: 1;
                border: 1px solid #eee;
                border-radius: 8px;
                padding: 20px;
                background-color: #fafafa;
                min-height: 400px;
            }
            .filter-panel {
                width: 300px;
                background-color: #e8f4fd;
                border: 1px solid #b3d9ff;
                border-radius: 8px;
                padding: 15px;
                height: fit-content;
            }
            .filter-panel h4 {
                margin: 0 0 15px 0;
                color: #333;
                font-size: 1.1em;
            }
            .filter-controls {
                display: flex;
                flex-direction: column;
                gap: 10px;
            }
            .filter-controls select, .filter-controls input, .filter-controls button {
                padding: 8px 12px;
                border: 1px solid #ddd;
                border-radius: 4px;
                font-size: 14px;
                width: 100%;
                box-sizing: border-box;
            }
            .filter-controls button {
                background-color: #667eea;
                color: white;
                border: none;
                cursor: pointer;
                transition: background-color 0.3s;
                margin-top: 10px;
            }
            .filter-controls button:hover {
                background-color: #5a6fd8;
            }
            .filter-controls button.reset {
                background-color: #dc3545;
            }
            .filter-controls button.reset:hover {
                background-color: #c82333;
            }
            .stats-grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 20px;
                margin: 20px 0;
            }
            .stat-card {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 20px;
                border-radius: 8px;
                text-align: center;
            }
            .stat-card h4 {
                margin: 0 0 10px 0;
                font-size: 1.1em;
                opacity: 0.9;
            }
            .stat-card .value {
                font-size: 2em;
                font-weight: bold;
                margin: 0;
            }
            .summary-table {
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
            }
            .summary-table th, .summary-table td {
                padding: 12px;
                text-align: left;
                border-bottom: 1px solid #ddd;
            }
            .summary-table th {
                background-color: #667eea;
                color: white;
                font-weight: bold;
            }
            .summary-table tr:nth-child(even) {
                background-color: #f9f9f9;
            }
            .summary-table tr:hover {
                background-color: #f0f0f0;
            }
            .best-performance {
                background-color: #ffd700 !important;
                font-weight: bold;
            }
            .no-data {
                text-align: center;
                padding: 40px;
                color: #666;
                font-style: italic;
            }
            .debug-info {
                background-color: #fff3cd;
                border: 1px solid #ffeaa7;
                border-radius: 4px;
                padding: 10px;
                margin: 10px 0;
                font-size: 0.9em;
                color: #856404;
            }
            .recommendations {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                border-radius: 8px;
                padding: 25px;
                margin: 20px 0;
            }
            .recommendations h3 {
                margin-top: 0;
                font-size: 1.5em;
                border-bottom: 2px solid rgba(255,255,255,0.3);
                padding-bottom: 10px;
            }
            .recommendations h4 {
                color: #ffd700;
                margin: 20px 0 10px 0;
                font-size: 1.2em;
            }
            .recommendations ul {
                margin: 10px 0;
                padding-left: 20px;
            }
            .recommendations li {
                margin: 8px 0;
                line-height: 1.6;
            }
            .criteria-section {
                background-color: #e8f4fd;
                border: 1px solid #b3d9ff;
                border-radius: 8px;
                padding: 20px;
                margin: 15px 0;
                color: #0f5132;
            }
            .criteria-section h4 {
                color: #0f5132;
                margin-top: 0;
            }
            .criteria-section li {
                margin: 8px 0;
                line-height: 1.5;
            }
            .db-info {
                background-color: #d4edda;
                border: 1px solid #c3e6cb;
                border-radius: 8px;
                padding: 15px;
                margin: 20px 0;
                color: #155724;
            }
            .db-info h4 {
                margin: 0 0 10px 0;
                color: #0f5132;
            }
            .footer {
                background-color: #333;
                color: white;
                text-align: center;
                padding: 20px;
                margin-top: 40px;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🎯 Optuna HPO 통합 DB 대시보드 (상관관계 개선)</h1>
                <p>all_studies.db 기반 | 개선된 파라미터 상관관계 차트 + 모든 기능</p>
                <p>생성 시간: """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """</p>
            </div>
            
            <div class="content">
    """
    
    # DB 정보 섹션
    html_content += f"""
                <div class="db-info">
                    <h4>📂 통합 DB 정보</h4>
                    <p><strong>DB 파일:</strong> optuna_studies/all_studies.db</p>
                    <p><strong>저장된 Study:</strong> """ + ", ".join([f"{model}_hpo_study" for model in studies.keys()]) + """</p>
                    <p><strong>총 Study 수:</strong> """ + str(len(studies)) + """개</p>
                    <p><strong>개선사항:</strong> 파라미터 상관관계 명확한 표시 + 강도별 색상 구분</p>
                </div>
    """
    
    # 전체 실험 요약 (기존과 동일)
    html_content += """
                <div class="section">
                    <h2>📊 전체 실험 요약</h2>
    """
    
    # 통계 계산
    total_trials = 0
    best_performances = {}
    for model_type, study in studies.items():
        trials = [t for t in study.trials if t.state == optuna.trial.TrialState.COMPLETE]
        total_trials += len(trials)
        if trials:
            best_performances[model_type] = study.best_value
    
    if best_performances:
        best_overall = max(best_performances.values())
        best_model = max(best_performances, key=best_performances.get)
        
        html_content += f"""
                    <div class="stats-grid">
                        <div class="stat-card">
                            <h4>총 Trial 수</h4>
                            <div class="value">{total_trials}</div>
                        </div>
                        <div class="stat-card">
                            <h4>최고 성능</h4>
                            <div class="value">{best_overall:.4f}</div>
                        </div>
                        <div class="stat-card">
                            <h4>최고 모델</h4>
                            <div class="value">{best_model}</div>
                        </div>
                        <div class="stat-card">
                            <h4>분석 모델 수</h4>
                            <div class="value">{len(studies)}</div>
                        </div>
                    </div>
        """
    
    # 모델별 성능 요약 테이블
    html_content += """
                    <h3>🏆 모델별 성능 요약</h3>
                    <table class="summary-table">
                        <thead>
                            <tr>
                                <th>모델</th>
                                <th>최고 성능</th>
                                <th>평균 성능</th>
                                <th>표준편차</th>
                                <th>성공률</th>
                                <th>수렴 상태</th>
                            </tr>
                        </thead>
                        <tbody>
    """
    
    for model_type, history in history_results.items():
        if history:
            row_class = "best-performance" if history['best_value'] == best_overall else ""
            html_content += f"""
                            <tr class="{row_class}">
                                <td><strong>{model_type}</strong></td>
                                <td>{history['best_value']:.4f}</td>
                                <td>{history['mean_value']:.4f}</td>
                                <td>{history['std_value']:.4f}</td>
                                <td>{history['success_rate']:.1f}%</td>
                                <td>{history['convergence']}</td>
                            </tr>
            """
    
    html_content += """
                        </tbody>
                    </table>
                </div>
    """
    
    # 각 모델별 완전한 분석 (기존 모든 차트 포함, 상관관계만 개선)
    for model_type, study in studies.items():
        model_data = all_data.get(model_type, {})
        if not model_data:
            continue
            
        html_content += f"""
                <div class="model-section">
                    <h3>🎯 {model_type} 완전한 분석 (통합 DB)</h3>
        """
        
        # 기본 통계
        history = history_results.get(model_type, {})
        if history:
            html_content += f"""
                    <div class="stats-grid">
                        <div class="stat-card">
                            <h4>최고 성능</h4>
                            <div class="value">{history['best_value']:.4f}</div>
                        </div>
                        <div class="stat-card">
                            <h4>평균 성능</h4>
                            <div class="value">{history['mean_value']:.4f}</div>
                        </div>
                        <div class="stat-card">
                            <h4>표준편차</h4>
                            <div class="value">{history['std_value']:.4f}</div>
                        </div>
                        <div class="stat-card">
                            <h4>성능 범위</h4>
                            <div class="value">{history['min_value']:.4f} ~ {history['max_value']:.4f}</div>
                        </div>
                    </div>
            """
        
        # 1. 최적화 과정
        html_content += f"""
                    <h4>📈 최적화 과정</h4>
                    <div class="chart-container">
                        <div id="optimization_{model_type}" style="height: 400px;"></div>
                    </div>
        """
        
        # 2. 파라미터 중요도
        html_content += f"""
                    <h4>🎯 파라미터 중요도</h4>
                    <div class="chart-container">
                        <div id="importance_{model_type}" style="height: 400px;"></div>
                    </div>
        """
        
        # 3. 파라미터 상관관계 (개선된 버전)
        html_content += f"""
                    <h4>🔗 파라미터 상관관계 (개선됨)</h4>
                    <div class="chart-container">
                        <div id="correlation_{model_type}" style="height: 500px;"></div>
                    </div>
        """
        
        # 추가 차트들 (Parallel Coordinate, Contour, Slice) - 기존과 동일
        html_content += f"""
                    <h4>🔄 Parallel Coordinate Plot</h4>
                    <div class="chart-row">
                        <div class="chart-container">
                            <div id="parallel_{model_type}" style="height: 400px;"></div>
                        </div>
                        <div class="filter-panel">
                            <h4>🎚️ 성능 범위 필터</h4>
                            <div class="filter-controls">
                                <label>최소값: <span id="{model_type.lower()}-parallel-min-value">0.50</span></label>
                                <input type="range" id="{model_type.lower()}-parallel-min" min="0" max="1" step="0.01" value="0.5">
                                <label>최대값: <span id="{model_type.lower()}-parallel-max-value">1.00</span></label>
                                <input type="range" id="{model_type.lower()}-parallel-max" min="0" max="1" step="0.01" value="1.0">
                                <button onclick="createParallelCoordinate('{model_type}')">필터 적용</button>
                                <button class="reset" onclick="resetParallelCoordinate('{model_type}')">리셋</button>
                            </div>
                        </div>
                    </div>
                    
                    <h4>🌐 Contour Plot</h4>
                    <div class="chart-row">
                        <div class="chart-container">
                            <div id="contour_{model_type}" style="height: 400px;"></div>
                        </div>
                        <div class="filter-panel">
                            <h4>📊 축 선택</h4>
                            <div class="filter-controls">
                                <label>X축 파라미터:</label>
                                <select id="{model_type.lower()}-contour-x">
        """
        
        for param in model_data.get('param_names', []):
            html_content += f'<option value="{param}">{param}</option>'
        
        html_content += f"""
                                </select>
                                <label>Y축 파라미터:</label>
                                <select id="{model_type.lower()}-contour-y">
        """
        
        for param in model_data.get('param_names', []):
            html_content += f'<option value="{param}">{param}</option>'
        
        html_content += f"""
                                </select>
                                <button onclick="createContourPlot('{model_type}')">차트 생성</button>
                                <button class="reset" onclick="resetContourPlot('{model_type}')">리셋</button>
                            </div>
                        </div>
                    </div>
                    
                    <h4>🔍 Slice Plot</h4>
                    <div class="chart-row">
                        <div class="chart-container">
                            <div id="slice_{model_type}" style="height: 400px;"></div>
                        </div>
                        <div class="filter-panel">
                            <h4>⚙️ 파라미터 선택</h4>
                            <div class="filter-controls">
                                <label>파라미터:</label>
                                <select id="{model_type.lower()}-slice-param">
        """
        
        for param in model_data.get('param_names', []):
            html_content += f'<option value="{param}">{param}</option>'
        
        html_content += f"""
                                </select>
                                <label>성능 임계값: <span id="{model_type.lower()}-slice-range-value">0.50</span></label>
                                <input type="range" id="{model_type.lower()}-slice-range" min="0" max="1" step="0.01" value="0.5">
                                <button onclick="createSlicePlot('{model_type}')">차트 생성</button>
                                <button class="reset" onclick="resetSlicePlot('{model_type}')">리셋</button>
                            </div>
                        </div>
                    </div>
                </div>
        """
    
    # 사용자 지정 권장사항 생성
    custom_recommendations = generate_custom_recommendations(studies, importance_results, history_results)
    
    # 사용자 지정 권장사항 섹션
    html_content += f"""
                <div class="recommendations">
                    <h3>💡 다음 실험을 위한 권장사항 (통합 DB 기반)</h3>
                    {custom_recommendations}
                </div>
    """
    
    # JavaScript 코드 (개선된 상관관계 차트 포함)
    html_content += """
            </div>
            
            <div class="footer">
                <p>🎯 Optuna HPO 통합 DB 대시보드 (상관관계 개선) | 생성 시간: """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """</p>
                <p>📂 DB 파일: optuna_studies/all_studies.db</p>
            </div>
        </div>
        
        <script>
            // 모든 데이터
            const allData = """ + safe_json_dumps(all_data) + """;
            
            // 차트 생성 함수들
            function createOptimizationChart(modelType) {
                const data = allData[modelType];
                if (!data || !data.trials) return;
                
                const trials = data.trials;
                const x = trials.map(trial => trial.trial_number);
                const y = trials.map(trial => trial.value);
                
                const trace = {
                    x: x,
                    y: y,
                    type: 'scatter',
                    mode: 'lines+markers',
                    name: 'Performance',
                    line: { color: '#667eea', width: 2 },
                    marker: { size: 6, color: '#667eea' }
                };
                
                const layout = {
                    title: `${modelType} 최적화 과정`,
                    xaxis: { title: 'Trial Number' },
                    yaxis: { title: 'Performance' },
                    height: 400
                };
                
                Plotly.newPlot('optimization_' + modelType, [trace], layout);
            }
            
            function createImportanceChart(modelType) {
                const data = allData[modelType];
                if (!data || !data.importance || Object.keys(data.importance).length === 0) {
                    document.getElementById('importance_' + modelType).innerHTML = '<div class="no-data">파라미터 중요도 데이터가 없습니다.</div>';
                    return;
                }
                
                const importance = data.importance;
                const sortedParams = Object.entries(importance).sort((a, b) => b[1] - a[1]);
                
                const trace = {
                    x: sortedParams.map(item => item[1]),
                    y: sortedParams.map(item => item[0]),
                    type: 'bar',
                    orientation: 'h',
                    marker: { color: '#667eea' }
                };
                
                const layout = {
                    title: `${modelType} 파라미터 중요도`,
                    xaxis: { title: 'Importance' },
                    yaxis: { title: 'Parameters' },
                    height: 400
                };
                
                Plotly.newPlot('importance_' + modelType, [trace], layout);
            }
            
            function createCorrelationChart(modelType) {
                const data = allData[modelType];
                if (!data || !data.correlation || Object.keys(data.correlation).length === 0) {
                    document.getElementById('correlation_' + modelType).innerHTML = '<div class="no-data">상관관계 데이터가 없습니다.</div>';
                    return;
                }
                
                const correlation = data.correlation;
                const sortedCorr = Object.entries(correlation).sort((a, b) => Math.abs(b[1]) - Math.abs(a[1]));
                
                const trace = {
                    x: sortedCorr.map(item => item[1]),
                    y: sortedCorr.map(item => item[0]),
                    type: 'bar',
                    orientation: 'h',
                    marker: { 
                        color: sortedCorr.map(item => {
                            const corr = item[1];
                            if (corr > 0.7) return '#d32f2f';      // 강한 양의 상관: 빨간색
                            else if (corr > 0.3) return '#f57c00'; // 중간 양의 상관: 주황색
                            else if (corr > 0) return '#388e3c';   // 약한 양의 상관: 녹색
                            else if (corr > -0.3) return '#1976d2'; // 약한 음의 상관: 파란색
                            else if (corr > -0.7) return '#7b1fa2'; // 중간 음의 상관: 보라색
                            else return '#424242';                  // 강한 음의 상관: 회색
                        })
                    },
                    text: sortedCorr.map(item => item[1].toFixed(3)),
                    textposition: 'auto',
                    hovertemplate: 
                        '<b>%{y}</b><br>' +
                        '상관관계: %{x:.4f}<br>' +
                        '<extra></extra>'
                };
                
                const layout = {
                    title: `${modelType} 파라미터 상관관계 (강도별 정렬)`,
                    xaxis: { 
                        title: 'Correlation Coefficient', 
                        range: [-1.1, 1.1],
                        zeroline: true,
                        zerolinecolor: '#000',
                        zerolinewidth: 2,
                        tickformat: '.3f'
                    },
                    yaxis: { 
                        title: 'Parameter Pairs',
                        automargin: true,
                        tickfont: {size: 11}
                    },
                    height: 500,
                    margin: {l: 250, r: 50, t: 80, b: 80},
                    plot_bgcolor: '#fafafa',
                    paper_bgcolor: 'white'
                };
                
                Plotly.newPlot('correlation_' + modelType, [trace], layout);
                console.log(`${modelType} 상관관계 차트 생성 완료:`, sortedCorr.length, '개 쌍');
            }
            
            function createParallelCoordinate(modelType) {
                const data = allData[modelType];
                if (!data || !data.trials) return;
                
                const minSlider = document.getElementById(modelType.toLowerCase() + '-parallel-min');
                const maxSlider = document.getElementById(modelType.toLowerCase() + '-parallel-max');
                const minValue = parseFloat(minSlider.value);
                const maxValue = parseFloat(maxSlider.value);
                
                document.getElementById(modelType.toLowerCase() + '-parallel-min-value').textContent = minValue.toFixed(2);
                document.getElementById(modelType.toLowerCase() + '-parallel-max-value').textContent = maxValue.toFixed(2);
                
                const filteredTrials = data.trials.filter(trial => 
                    trial.value >= minValue && trial.value <= maxValue
                );
                
                if (filteredTrials.length === 0) {
                    document.getElementById('parallel_' + modelType).innerHTML = '<div class="no-data">필터 조건에 맞는 데이터가 없습니다.</div>';
                    return;
                }
                
                const paramNames = data.param_names || [];
                const dimensions = paramNames.map(param => ({
                    label: param,
                    values: filteredTrials.map(trial => trial[param] || 0)
                }));
                
                dimensions.push({
                    label: 'Performance',
                    values: filteredTrials.map(trial => trial.value)
                });
                
                const trace = {
                    type: 'parcoords',
                    dimensions: dimensions,
                    line: {
                        color: filteredTrials.map(trial => trial.value),
                        colorscale: 'Viridis',
                        showscale: true
                    }
                };
                
                const layout = {
                    title: `${modelType} Parallel Coordinate Plot`,
                    height: 400
                };
                
                Plotly.newPlot('parallel_' + modelType, [trace], layout);
            }
            
            function createContourPlot(modelType) {
                const data = allData[modelType];
                if (!data || !data.trials) return;
                
                const xParamSelect = document.getElementById(modelType.toLowerCase() + '-contour-x');
                const yParamSelect = document.getElementById(modelType.toLowerCase() + '-contour-y');
                const xParam = xParamSelect.value;
                const yParam = yParamSelect.value;
                
                if (!xParam || !yParam || xParam === yParam) {
                    document.getElementById('contour_' + modelType).innerHTML = '<div class="no-data">서로 다른 두 파라미터를 선택하세요.</div>';
                    return;
                }
                
                const validTrials = data.trials.filter(trial => 
                    trial[xParam] !== undefined && trial[yParam] !== undefined
                );
                
                if (validTrials.length < 3) {
                    document.getElementById('contour_' + modelType).innerHTML = '<div class="no-data">데이터가 부족합니다.</div>';
                    return;
                }
                
                const trace = {
                    x: validTrials.map(trial => trial[xParam]),
                    y: validTrials.map(trial => trial[yParam]),
                    z: validTrials.map(trial => trial.value),
                    type: 'scatter',
                    mode: 'markers',
                    marker: {
                        size: 10,
                        color: validTrials.map(trial => trial.value),
                        colorscale: 'Viridis',
                        showscale: true
                    }
                };
                
                const layout = {
                    title: `${modelType} Contour Plot (${xParam} vs ${yParam})`,
                    xaxis: { title: xParam },
                    yaxis: { title: yParam },
                    height: 400
                };
                
                Plotly.newPlot('contour_' + modelType, [trace], layout);
            }
            
            function createSlicePlot(modelType) {
                const data = allData[modelType];
                if (!data || !data.trials) return;
                
                const paramSelect = document.getElementById(modelType.toLowerCase() + '-slice-param');
                const rangeSlider = document.getElementById(modelType.toLowerCase() + '-slice-range');
                const param = paramSelect.value;
                const threshold = parseFloat(rangeSlider.value);
                
                document.getElementById(modelType.toLowerCase() + '-slice-range-value').textContent = threshold.toFixed(2);
                
                if (!param) {
                    document.getElementById('slice_' + modelType).innerHTML = '<div class="no-data">파라미터를 선택하세요.</div>';
                    return;
                }
                
                const filteredTrials = data.trials.filter(trial => 
                    trial[param] !== undefined && trial.value >= threshold
                );
                
                if (filteredTrials.length === 0) {
                    document.getElementById('slice_' + modelType).innerHTML = '<div class="no-data">조건에 맞는 데이터가 없습니다.</div>';
                    return;
                }
                
                const trace = {
                    x: filteredTrials.map(trial => trial[param]),
                    y: filteredTrials.map(trial => trial.value),
                    type: 'scatter',
                    mode: 'markers',
                    marker: {
                        size: 10,
                        color: filteredTrials.map(trial => trial.value),
                        colorscale: 'Viridis',
                        showscale: true
                    }
                };
                
                const layout = {
                    title: `${modelType} Slice Plot (${param})`,
                    xaxis: { title: param },
                    yaxis: { title: 'Performance' },
                    height: 400
                };
                
                Plotly.newPlot('slice_' + modelType, [trace], layout);
            }
            
            // 리셋 함수들
            function resetParallelCoordinate(modelType) {
                document.getElementById(modelType.toLowerCase() + '-parallel-min').value = 0.5;
                document.getElementById(modelType.toLowerCase() + '-parallel-max').value = 1.0;
                document.getElementById(modelType.toLowerCase() + '-parallel-min-value').textContent = '0.50';
                document.getElementById(modelType.toLowerCase() + '-parallel-max-value').textContent = '1.00';
                createParallelCoordinate(modelType);
            }
            
            function resetContourPlot(modelType) {
                createContourPlot(modelType);
            }
            
            function resetSlicePlot(modelType) {
                document.getElementById(modelType.toLowerCase() + '-slice-range').value = 0.5;
                document.getElementById(modelType.toLowerCase() + '-slice-range-value').textContent = '0.50';
                createSlicePlot(modelType);
            }
            
            // 페이지 로드 완료 후 모든 차트 초기화
            window.addEventListener('load', function() {
                console.log('통합 DB 대시보드 (상관관계 개선) 로드 완료');
                
                // 각 모델에 대해 모든 차트 생성
                Object.keys(allData).forEach(modelType => {
                    createOptimizationChart(modelType);
                    createImportanceChart(modelType);
                    createCorrelationChart(modelType);
                    createParallelCoordinate(modelType);
                    createContourPlot(modelType);
                    createSlicePlot(modelType);
                });
            });
        </script>
    </body>
    </html>
    """
    
    # 실험별 결과 폴더 생성
    experiment_results_dir = f"results/{experiment_name}"
    os.makedirs(experiment_results_dir, exist_ok=True)
    
    # HTML 파일 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{experiment_results_dir}/optuna_unified_dashboard_correlation_fixed_{experiment_name}_{timestamp}.html"
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"✅ 상관관계 개선 대시보드가 '{filename}'에 저장되었습니다!")
    return filename

def get_model_types_from_db(db_path):
    """DB에서 실제 study_name들을 확인합니다."""
    try:
        conn = sqlite3.connect(db_path)
        
        # 실제 study_name 확인
        query = "SELECT DISTINCT study_name FROM studies"
        studies_df = pd.read_sql_query(query, conn)
        
        print("🔍 DB에 있는 실제 study_name들:")
        for study_name in studies_df['study_name']:
            print(f"   - {study_name}")
        
        conn.close()
        return studies_df['study_name'].tolist()
        
    except Exception as e:
        print(f"❌ DB 읽기 오류: {e}")
        return []

def run_optuna_dashboard(db_path):
    """Optuna Dashboard를 터미널 명령어로 실행합니다."""
    try:
        # DB 경로
        # db_path = "optuna_studies/titanic_5models_hpo_v1.db" # 이 부분을 동적으로 변경
        
        print("🚀 Optuna Dashboard 시작...")
        print(f"📊 DB: {db_path}")
        
        # optuna-dashboard 명령어 실행
        cmd = f"optuna-dashboard sqlite:///{db_path}"
        print(f"💻 실행: {cmd}")
        
        subprocess.run(cmd, shell=True, check=True)
        
    except subprocess.CalledProcessError:
        print("⚠️  수동 실행 필요:")
        print(f"   optuna-dashboard sqlite:///{db_path}")
    except Exception as e:
        print(f"❌ 오류: {e}")

if __name__ == "__main__":
    import sys
    # 명령행 인자 처리
    if len(sys.argv) > 1:
        experiment_name = sys.argv[1]
    else:
        experiment_name = "titanic_5models_hpo_v1"  # 기본값
    
    print(f"🎯 실험 이름: {experiment_name}")
    
    # 올바른 DB 경로 설정
    db_path = f"optuna_studies/{experiment_name}/all_studies.db"
    
    print(f"🔍 DB 경로: {db_path}")
    
    # DB에서 실제 study_name들 읽기
    actual_study_names = get_model_types_from_db(db_path)
    
    if not actual_study_names:
        print("❌ 사용 가능한 study가 없습니다!")
        exit() # 프로그램 종료
    
    # 각 study별로 로드
    studies = {}
    for study_name in actual_study_names:
        try:
            study = optuna.load_study(
                study_name=study_name,  # 실제 study_name 사용
                storage=f'sqlite:///{db_path}'
            )
            studies[study_name] = study
            print(f"✅ {study_name} study 로드 완료")
        except Exception as e:
            print(f"❌ {study_name} study 로드 실패: {e}")
    
    if not studies:
        print("❌ 로드할 study가 없습니다!")
        exit() # 프로그램 종료
    
    # 분석 수행
    importance_results = analyze_parameter_importance(studies)
    history_results = analyze_optimization_history(studies)
    
    # 1. HTML 대시보드 생성 (상관관계 차트 개선)
    html_file = create_unified_db_dashboard_with_fixed_correlation(studies, experiment_name)
    
    # 2. 엑셀 보고서 생성 (기존 고급 서식과 동일)
    excel_file = create_excel_report_unified_db(studies, importance_results, history_results, experiment_name)
    
    run_optuna_dashboard(db_path)
    print("\n🎉 상관관계 개선 완료!")
    print("📋 HTML 대시보드 개선사항:")
    print("  ✅ 파라미터 상관관계: '파라미터1 ↔ 파라미터2' 형식으로 명확한 표시")
    print("  ✅ 중복 제거: 상삼각 매트릭스만 표시 (A-B와 B-A 중복 제거)")
    print("  ✅ 강도별 색상: 상관관계 강도에 따른 6단계 색상 구분")
    print("  ✅ 정렬: 강한 상관관계부터 약한 상관관계 순으로 정렬")
    print("  ✅ 레이아웃: 충분한 여백과 적절한 폰트 크기로 가독성 향상")
    print("  ✅ 기존 기능: 모든 차트 + 필터 + 사용자 지정 권장사항 유지")
    print("\n📊 엑셀 보고서:")
    print("  ✅ 기존 고급 서식과 동일한 구조 및 조건부 서식")
    print(f"\n📂 생성된 파일:")
    print(f"  HTML: {html_file}")
    print(f"  Excel: {excel_file}")
    print("🌐 웹 브라우저에서 HTML을 열어서 개선된 상관관계 차트를 확인하세요!")
    print("📊 캡처하신 문제가 해결되었는지 확인해보세요!") 