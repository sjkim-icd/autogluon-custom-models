import json
import pandas as pd
import numpy as np
import optuna
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference

def load_studies():
    """모든 Optuna study 로드"""
    model_types = ['DCNV2', 'CUSTOM_FOCAL_DL', 'RF']
    studies = {}
    
    for model_type in model_types:
        try:
            study = optuna.load_study(
                study_name=f'{model_type}_hpo_study',
                storage=f'sqlite:///optuna_studies/{model_type}_study.db'
            )
            studies[model_type] = study
            print(f"✅ {model_type} study 로드 완료")
        except Exception as e:
            print(f"❌ {model_type} study 로드 실패: {e}")
    
    return studies

def analyze_parameter_importance(studies):
    """파라미터 중요도 분석"""
    importance_results = {}
    
    for model_type, study in studies.items():
        try:
            importance = optuna.importance.get_param_importances(study)
            importance_results[model_type] = importance
            print(f"✅ {model_type} 파라미터 중요도 계산 완료")
        except Exception as e:
            print(f"❌ {model_type} 파라미터 중요도 계산 실패: {e}")
    
    return importance_results

def analyze_optimization_history(studies):
    """최적화 과정 분석"""
    history_results = {}
    
    for model_type, study in studies.items():
        try:
            trials = study.trials
            successful_trials = [t for t in trials if t.state == optuna.trial.TrialState.COMPLETE]
            
            if len(successful_trials) >= 2:
                values = [t.value for t in successful_trials]
                best_value = study.best_value
                mean_value = np.mean(values)
                std_value = np.std(values)
                
                # 수렴성 분석
                if len(values) >= 5:
                    recent_values = values[-5:]
                    improvement = recent_values[-1] - recent_values[0]
                    
                    if improvement > 0.01:
                        convergence = "개선 중"
                    elif abs(improvement) < 0.005:
                        convergence = "수렴됨"
                    else:
                        convergence = "불안정"
                else:
                    convergence = "데이터 부족"
                
                history_results[model_type] = {
                    'best_value': best_value,
                    'mean_value': mean_value,
                    'std_value': std_value,
                    'min_value': min(values),
                    'max_value': max(values),
                    'convergence': convergence,
                    'success_rate': len(successful_trials)/len(trials)*100
                }
                
        except Exception as e:
            print(f"❌ {model_type} 최적화 과정 분석 실패: {e}")
    
    return history_results

def create_excel_report(studies, importance_results, history_results):
    """고급 분석 리포트 엑셀 생성"""
    print("=== 고급 분석 리포트 엑셀 생성 ===")
    
    # 워크북 생성
    wb = Workbook()
    
    # 기본 시트 제거
    wb.remove(wb.active)
    
    # 1. 요약 시트
    create_summary_sheet(wb, studies, history_results)
    
    # 2. 파라미터 중요도 시트
    create_importance_sheet(wb, importance_results)
    
    # 3. 최적화 과정 시트
    create_optimization_sheet(wb, history_results)
    
    # 4. 각 모델별 상세 시트
    for model_type in studies.keys():
        create_model_detail_sheet(wb, model_type, studies[model_type], 
                                importance_results.get(model_type, {}),
                                history_results.get(model_type, {}))
    
    # 5. 권장사항 시트
    create_recommendations_sheet(wb, studies, importance_results, history_results)
    
    # 파일 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"optuna_advanced_report_{timestamp}.xlsx"
    wb.save(filename)
    
    print(f"✅ 고급 분석 리포트가 '{filename}'에 저장되었습니다!")
    return filename

def create_summary_sheet(wb, studies, history_results):
    """요약 시트 생성"""
    ws = wb.create_sheet("📊 실험 요약")
    
    # 제목
    ws['A1'] = "🎯 Optuna HPO 고급 분석 리포트"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    # 생성 시간
    ws['A2'] = f"생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(size=10, italic=True)
    
    # 요약 통계
    row = 4
    ws[f'A{row}'] = "📈 전체 실험 통계"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row += 1
    total_trials = sum([len([t for t in studies[model].trials if t.state == optuna.trial.TrialState.COMPLETE]) 
                       for model in studies.keys()])
    best_performance = max([history.get('best_value', 0) for history in history_results.values()])
    avg_performance = np.mean([history.get('mean_value', 0) for history in history_results.values()])
    
    summary_data = [
        ["총 Trial 수", total_trials],
        ["최고 성능", f"{best_performance:.4f}"],
        ["평균 성능", f"{avg_performance:.4f}"],
        ["분석 모델 수", len(studies)]
    ]
    
    for i, (label, value) in enumerate(summary_data):
        ws[f'A{row+i}'] = label
        ws[f'B{row+i}'] = value
        ws[f'A{row+i}'].font = Font(bold=True)
    
    # 모델별 성능 요약
    row += 6
    ws[f'A{row}'] = "🏆 모델별 성능 요약"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row += 1
    headers = ["모델", "최고 성능", "평균 성능", "표준편차", "수렴 상태", "성공률"]
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=i+1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    row += 1
    for model_type, history in history_results.items():
        ws[f'A{row}'] = model_type
        ws[f'B{row}'] = history['best_value']
        ws[f'C{row}'] = history['mean_value']
        ws[f'D{row}'] = history['std_value']
        ws[f'E{row}'] = history['convergence']
        ws[f'F{row}'] = f"{history['success_rate']:.1f}%"
        
        # 조건부 서식: 최고 성능 강조
        if history['best_value'] == best_performance:
            ws[f'B{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        row += 1

def create_importance_sheet(wb, importance_results):
    """파라미터 중요도 시트 생성"""
    ws = wb.create_sheet("🔍 파라미터 중요도")
    
    # 제목
    ws['A1'] = "파라미터 중요도 분석"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:C1')
    
    row = 3
    for model_type, importance in importance_results.items():
        ws[f'A{row}'] = f"📊 {model_type}"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row += 1
        headers = ["파라미터", "중요도", "순위"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row += 1
        sorted_importance = sorted(importance.items(), key=lambda x: x[1], reverse=True)
        for rank, (param, score) in enumerate(sorted_importance, 1):
            ws[f'A{row}'] = param
            ws[f'B{row}'] = score
            ws[f'C{row}'] = rank
            
            # 상위 3개 강조
            if rank <= 3:
                ws[f'A{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                ws[f'B{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                ws[f'C{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            row += 1
        
        row += 2

def create_optimization_sheet(wb, history_results):
    """최적화 과정 시트 생성"""
    ws = wb.create_sheet("📈 최적화 과정")
    
    # 제목
    ws['A1'] = "최적화 과정 분석"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    row = 3
    headers = ["모델", "최고 성능", "평균 성능", "표준편차", "성능 범위", "수렴 상태"]
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=i+1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    row += 1
    for model_type, history in history_results.items():
        ws[f'A{row}'] = model_type
        ws[f'B{row}'] = history['best_value']
        ws[f'C{row}'] = history['mean_value']
        ws[f'D{row}'] = history['std_value']
        ws[f'E{row}'] = f"{history['min_value']:.4f} ~ {history['max_value']:.4f}"
        ws[f'F{row}'] = history['convergence']
        
        # 조건부 서식
        if history['convergence'] == "수렴됨":
            ws[f'F{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif history['convergence'] == "불안정":
            ws[f'F{row}'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        row += 1

def create_model_detail_sheet(wb, model_type, study, importance, history):
    """모델별 상세 시트 생성"""
    ws = wb.create_sheet(f"🎯 {model_type}")
    
    # 제목
    ws['A1'] = f"{model_type} 상세 분석"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:D1')
    
    # 기본 정보
    row = 3
    ws[f'A{row}'] = "📊 기본 정보"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row += 1
    if history:
        basic_info = [
            ["최고 성능", f"{history['best_value']:.4f}"],
            ["평균 성능", f"{history['mean_value']:.4f}"],
            ["표준편차", f"{history['std_value']:.4f}"],
            ["성능 범위", f"{history['min_value']:.4f} ~ {history['max_value']:.4f}"],
            ["수렴 상태", history['convergence']],
            ["성공률", f"{history['success_rate']:.1f}%"]
        ]
        
        for i, (label, value) in enumerate(basic_info):
            ws[f'A{row+i}'] = label
            ws[f'B{row+i}'] = value
            ws[f'A{row+i}'].font = Font(bold=True)
    
    # 파라미터 중요도
    row += 8
    ws[f'A{row}'] = "🔍 파라미터 중요도"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row += 1
    headers = ["파라미터", "중요도", "순위"]
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=i+1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    row += 1
    if importance:
        sorted_importance = sorted(importance.items(), key=lambda x: x[1], reverse=True)
        for rank, (param, score) in enumerate(sorted_importance, 1):
            ws[f'A{row}'] = param
            ws[f'B{row}'] = score
            ws[f'C{row}'] = rank
            
            # 상위 3개 강조
            if rank <= 3:
                ws[f'A{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                ws[f'B{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                ws[f'C{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            row += 1

def create_recommendations_sheet(wb, studies, importance_results, history_results):
    """권장사항 시트 생성"""
    ws = wb.create_sheet("💡 권장사항")
    
    # 제목
    ws['A1'] = "다음 실험을 위한 권장사항"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:C1')
    
    row = 3
    for model_type in studies.keys():
        ws[f'A{row}'] = f"🎯 {model_type} 권장사항"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row += 1
        
        # 파라미터 중요도 기반 권장사항
        if model_type in importance_results:
            importance = importance_results[model_type]
            top_params = sorted(importance.items(), key=lambda x: x[1], reverse=True)[:3]
            
            ws[f'A{row}'] = "📊 가장 중요한 파라미터:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for param, score in top_params:
                ws[f'A{row}'] = f"  • {param}: {score:.4f}"
                row += 1
        
        # 최적화 과정 기반 권장사항
        if model_type in history_results:
            history = history_results[model_type]
            
            ws[f'A{row}'] = "📈 최적화 상태:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = f"  • 최고 성능: {history['best_value']:.4f}"
            row += 1
            ws[f'A{row}'] = f"  • 성능 변동성: {history['std_value']:.4f}"
            row += 1
            ws[f'A{row}'] = f"  • 수렴 상태: {history['convergence']}"
            row += 1
            
            # 권장사항
            ws[f'A{row}'] = "💡 권장사항:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            if history['std_value'] > 0.05:
                ws[f'A{row}'] = "  • 더 많은 trial 필요 (높은 변동성)"
            elif history['convergence'] == "수렴됨":
                ws[f'A{row}'] = "  • 현재 설정으로 충분히 최적화됨"
            else:
                ws[f'A{row}'] = "  • 더 세밀한 파라미터 탐색 필요"
            
            row += 1
        
        row += 2

if __name__ == "__main__":
    # 모든 study 로드
    studies = load_studies()
    
    if not studies:
        print("❌ 로드할 study가 없습니다!")
    else:
        # 파라미터 중요도 분석
        importance_results = analyze_parameter_importance(studies)
        
        # 최적화 과정 분석
        history_results = analyze_optimization_history(studies)
        
        # 엑셀 리포트 생성
        report_file = create_excel_report(studies, importance_results, history_results)
        
        print("\n🎉 고급 분석 리포트 엑셀 생성 완료!")
        print("📋 리포트 내용:")
        print("  - 실험 요약 및 통계")
        print("  - 파라미터 중요도 분석")
        print("  - 최적화 과정 분석")
        print("  - 각 모델별 상세 분석")
        print("  - 다음 실험을 위한 권장사항")
        print(f"\n📂 파일 위치: {report_file}")
        print("📊 조건부 서식이 적용된 깔끔한 엑셀 리포트입니다!") 